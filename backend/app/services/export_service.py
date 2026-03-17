"""리포트 내보내기 서비스 — Excel/CSV 생성"""
import io
import csv
import logging
from datetime import date
from typing import Optional
from uuid import UUID

from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.invoice import Invoice
from app.models.vendor import Vendor
from app.models.invoice_payment import InvoicePayment

logger = logging.getLogger(__name__)


async def export_invoice_list(
    db: AsyncSession,
    company_id: UUID,
    status_filter: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    fmt: str = "csv",
) -> tuple[bytes, str, str]:
    """인보이스 목록 내보내기 → (content, filename, content_type)"""
    query = (
        select(Invoice, Vendor.company_name.label("vendor_name"))
        .outerjoin(Vendor, Invoice.vendor_id == Vendor.id)
        .where(Invoice.company_id == company_id)
    )
    if status_filter:
        query = query.where(Invoice.status == status_filter)
    if date_from:
        query = query.where(Invoice.invoice_date >= date_from)
    if date_to:
        query = query.where(Invoice.invoice_date <= date_to)

    query = query.order_by(Invoice.created_at.desc())
    result = await db.execute(query)

    headers = [
        "Invoice Number", "Vendor", "Invoice Date", "Due Date",
        "Subtotal", "Tax", "Total", "Currency", "Status",
        "Validation", "Source", "Created At",
    ]

    rows = []
    for invoice, vendor_name in result.all():
        rows.append([
            invoice.invoice_number or "",
            vendor_name or "",
            str(invoice.invoice_date) if invoice.invoice_date else "",
            str(invoice.due_date) if invoice.due_date else "",
            float(invoice.amount_subtotal),
            float(invoice.amount_tax),
            float(invoice.amount_total),
            invoice.currency_original,
            invoice.status,
            invoice.validation_status or "",
            invoice.source_channel,
            invoice.created_at.strftime("%Y-%m-%d %H:%M") if invoice.created_at else "",
        ])

    if fmt == "excel":
        return _to_excel(headers, rows), "invoices.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return _to_csv(headers, rows), "invoices.csv", "text/csv"


async def export_vendor_spend(
    db: AsyncSession,
    company_id: UUID,
    fmt: str = "csv",
) -> tuple[bytes, str, str]:
    """벤더별 지출 요약 내보내기"""
    result = await db.execute(
        select(
            Vendor.vendor_code,
            Vendor.company_name,
            Vendor.is_1099_required,
            func.count(Invoice.id).label("invoice_count"),
            func.coalesce(func.sum(Invoice.amount_total), 0).label("total_spend"),
        )
        .outerjoin(Invoice, Vendor.id == Invoice.vendor_id)
        .where(Vendor.company_id == company_id)
        .group_by(Vendor.vendor_code, Vendor.company_name, Vendor.is_1099_required)
        .order_by(func.sum(Invoice.amount_total).desc().nulls_last())
    )

    headers = ["Vendor Code", "Vendor Name", "1099 Required", "Invoice Count", "Total Spend"]
    rows = [
        [row.vendor_code, row.company_name, "Yes" if row.is_1099_required else "No",
         row.invoice_count, float(row.total_spend)]
        for row in result.all()
    ]

    if fmt == "excel":
        return _to_excel(headers, rows), "vendor_spend.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return _to_csv(headers, rows), "vendor_spend.csv", "text/csv"


async def export_payment_report(
    db: AsyncSession,
    company_id: UUID,
    fmt: str = "csv",
) -> tuple[bytes, str, str]:
    """결제 리포트 내보내기"""
    result = await db.execute(
        select(InvoicePayment, Invoice.invoice_number)
        .join(Invoice, InvoicePayment.invoice_id == Invoice.id)
        .where(InvoicePayment.company_id == company_id)
        .order_by(InvoicePayment.created_at.desc())
    )

    headers = [
        "Invoice Number", "Payment Method", "Status", "Scheduled Date",
        "Paid Date", "Amount", "Transaction Ref", "Bank", "Notes",
    ]
    rows = [
        [
            inv_num or "",
            p.payment_method, p.payment_status,
            str(p.scheduled_date) if p.scheduled_date else "",
            str(p.paid_date) if p.paid_date else "",
            float(p.amount_paid),
            p.transaction_ref or "",
            p.bank_name or "",
            p.notes or "",
        ]
        for p, inv_num in result.all()
    ]

    if fmt == "excel":
        return _to_excel(headers, rows), "payments.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return _to_csv(headers, rows), "payments.csv", "text/csv"


async def export_1099_prep(
    db: AsyncSession,
    company_id: UUID,
    tax_year: int,
    fmt: str = "csv",
) -> tuple[bytes, str, str]:
    """1099 준비 리포트 — 1099 대상 벤더의 연간 지출"""
    year_start = date(tax_year, 1, 1)
    year_end = date(tax_year, 12, 31)

    result = await db.execute(
        select(
            Vendor.vendor_code,
            Vendor.company_name,
            Vendor.ein,
            func.coalesce(func.sum(Invoice.amount_total), 0).label("annual_spend"),
            func.count(Invoice.id).label("invoice_count"),
        )
        .outerjoin(
            Invoice,
            (Vendor.id == Invoice.vendor_id) &
            (Invoice.invoice_date >= year_start) &
            (Invoice.invoice_date <= year_end) &
            (Invoice.status.in_(["APPROVED", "PAID", "SCHEDULED"]))
        )
        .where(
            Vendor.company_id == company_id,
            Vendor.is_1099_required == True,
        )
        .group_by(Vendor.vendor_code, Vendor.company_name, Vendor.ein)
        .order_by(func.sum(Invoice.amount_total).desc().nulls_last())
    )

    headers = ["Vendor Code", "Vendor Name", "EIN", "Annual Spend", "Invoice Count"]
    rows = [
        [row.vendor_code, row.company_name, row.ein or "", float(row.annual_spend), row.invoice_count]
        for row in result.all()
    ]

    filename = f"1099_prep_{tax_year}"
    if fmt == "excel":
        return _to_excel(headers, rows), f"{filename}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return _to_csv(headers, rows), f"{filename}.csv", "text/csv"


def _to_csv(headers: list[str], rows: list[list]) -> bytes:
    """CSV 바이트 생성"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")  # BOM for Excel 호환


def _to_excel(headers: list[str], rows: list[list]) -> bytes:
    """Excel 바이트 생성 (openpyxl)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active

        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 컬럼 너비 자동 조정
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    except ImportError:
        logger.warning("openpyxl not installed, falling back to CSV")
        return _to_csv(headers, rows)
